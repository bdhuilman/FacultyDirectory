import requests  # to be able to open web pages
import openpyxl  # to read and create spreadsheets
from openpyxl.styles import Font  # allow font size and type in the sheet to change
from bs4 import BeautifulSoup  # to parse the web page
import time  # to pause a bit after each page
from datetime import datetime  # to date the sheet name

# base URL for MCTC online directory; just add first+last
base_url = 'https://www.minneapolis.edu/employee-directory?field_department_target_id=All&title='
existing_spreadsheet = 'MinneapolisList.xlsx'  # spreadsheet of faculty
new_spreadsheet = 'MCTC_MSCF_Directory.xlsx'  # new spreadsheet of the directory
wb = openpyxl.load_workbook(existing_spreadsheet)  # open the existing faculty sheet
sheet = wb.active  # get the main (first) sheet
new_wb = openpyxl.Workbook()  # create a brand new spreadsheet in memory
new_sheet= new_wb.active  # set the main (first) sheet
now = datetime.now()  # now is a datetime object that represents the current date/time
# More info here: https://docs.python.org/3/library/datetime.html#strftime-strptime-behavior
now_string = now.strftime('%Y-%m-%d %H%M')  # strftime formats dates into strings.
new_sheet.title = 'MSCF Directory ' + now_string  # title the main sheet
# setup a dictionary of column numbers for each column header
d = {'First': 1, 'Last': 2, 'Dept': 3, 'Office': 4, 'Email': 5, 'Phone': 6}
new_sheet.cell(1,d['First'],'First')  # print a header row in the new sheet of the directory
new_sheet.cell(1,d['Last'],'Last')
new_sheet.cell(1,d['Dept'],'Dept')
new_sheet.cell(1,d['Office'],'Office')
new_sheet.cell(1,d['Email'],'Email')
new_sheet.cell(1,d['Phone'],'Phone')
boldFont = Font(size=14, bold=True)  # make a 14 pt. bold font
for i in range(6):  # first 6 columns; note that chr(65) is 'A', chr(66) is 'B', etc.
    new_sheet[f'{chr(65+i)}1'].font = boldFont  # apply the font to the header range
new_sheet.row_dimensions[1].height = 14  # set row size accordingly
new_sheet.freeze_panes = 'A2'  # freezes row 1
new_row = 2  # skip the header line
# start reading the existing faculty sheet
for row in range(2, sheet.max_row + 1):  # rows and columns start at 1, skip the header
    row_s = str(row)  # for use in col_letter+row_# format
    first_name = sheet['B' + row_s].value  # get relevant data from existing faculty sheet
    last_name = sheet['C' + row_s].value
    member = sheet['D' + row_s].value
    # print(f'{row}: {first_name} {last_name} - {member}')
    if member:  # if faculty is a member, go look them up online
        new_sheet.cell(new_row,d['First'],first_name)  # add name to the new directory sheet
        new_sheet.cell(new_row,d['Last'],last_name)
        print(f'{new_row}: {first_name} {last_name} : ', end='')  # status message on console
        url = base_url + f'{first_name}+{last_name}'  # build the full URL
        try:
            result = requests.get(url)  # try to open the URL
        except:
            print('You might not have Internet access.')  # fail gracefully
            exit(-1)
        if result.status_code != 200:  # ensure we have a legitimate response
            print('Error retrieving page.')  # fail gracefully otherwise
            exit(-2)
        src = result.content  # get the source of the page for BS
        soup = BeautifulSoup(src, 'lxml')  # use BS to get the content in a soup object
        tbody = soup.find_all('tbody')  # use BS to get the tbody element we want
        if len(tbody):  # if we got a result (a 0 evaluates to False)
            table_rows = tbody[0].find_all('tr')  # pull all the rows <tr> from the table
            for tr in table_rows:  # look at each <tr> in turn
                td = tr.find_all('td')  # pull all data <td> from each row
                data = [i.text for i in td]  # make a list of all the data
                # get what we need from the data, strip whitespace
                name = data[0].strip()  # pull the raw data from the web page
                dept = data[1].strip()
                office = data[2].strip()
                contact = data[3].strip()
                # the name field has extra junk and possibly many names in the <tr>
                # let's search for just the name we want in this row; returns -1 if not found
                if name.rfind(f'{first_name} {last_name}') >= 0:
                    print(f'{dept} : {office} : ', end='')
                    new_sheet.cell(new_row, d['Dept'], dept)  # add data to the new directory sheet
                    new_sheet.cell(new_row, d['Office'], office)
                    #print(f'{contact}')
                    c_list = contact.split('E: ')
                    email = c_list[1].split('P: ')[0]
                    try:  # some people don't have a phone number
                        phone = c_list[1].split('P: ')[1]
                    except IndexError:
                        phone = 'none'
                    print(f'{email} : {phone}', end='')
                    new_sheet.cell(new_row, d['Email'], email)  # add data to the new directory sheet
                    new_sheet.cell(new_row, d['Phone'], phone)
            print()  # close out the console line
            new_row += 1  # bump to the next row in the directory sheet
            new_wb.save(new_spreadsheet)  # save what we've got so far
            # time.sleep(2)  # let the web not think we're spamming
